import pyodbc
import requests
import json
import random
from urllib3.exceptions import InsecureRequestWarning
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from datetime import datetime, timedelta
import os


job_num = None
site_ID = None
equip_ID = None
Customer_ID = None
customer_Myid = None


filtered_row = None
highest_stock_item_id = float('-inf')  # Start with the lowest possible value
stock_code = None
stock_description = None
stock_quantity = None
stock_status = None
stock_serial = None

error_message = None

def check_customer_exists(customer_name):
    try:
        url = "https://apis.synchroteam.com/api/v3/customer/details"
        querystring = {"name": customer_name}
        headers = {
            'authorization': "Basic ****************=",
            'accept': "text/json",
            'content-type': "application/json",
            'cache-control': "no-cache"
        }
        response = requests.get(url, headers=headers, params=querystring)
        if response.status_code == 200:
            data = response.json()
            if 'id' in data:
                return data['id']  # Return the customer ID
        return None  # Customer does not exist in Synchroteam or an error occurred
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")

def check_site_exists(site_name):
    try:
        global site_ID
        global Customer_ID
        url = "https://apis.synchroteam.com/api/v3/site/details"
        querystring = {"name": site_name}
        headers = {
            'authorization': "Basic *****************=",
            'accept': "text/json",
            'content-type': "application/json",
            'cache-control': "no-cache"
        }
        response = requests.get(url, headers=headers, params=querystring)
        if response.status_code == 200:
            data = json.loads(response.text)
            site_ID = data["id"]
            Customer_ID = data["customer"]["id"]
            return True  # Site exists in Synchroteam
        else:
            return False  # Site does not exist in Synchroteam or an error occurred
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")

def check_equipment_exists(equipment_name):
    try:
        global equip_ID
        url = "https://apis.synchroteam.com/api/v3/equipment/details"
        querystring = {"name": equipment_name}
        headers = {
            'authorization': "Basic **********************=",
            'accept': "text/json",
            'content-type': "application/json",
            'cache-control': "no-cache"
        }
        response = requests.get(url, headers=headers, params=querystring)
        if response.status_code == 200:
            data = json.loads(response.text)
            equip_ID = data["id"]
            return True  # Equipment exists in Synchroteam
        else:
            return False  # Equipment does not exist in Synchroteam or an error occurred
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")

def check_if_job_exists(job_number):
    try:
        global job_num
        url = "https://apis.synchroteam.com/api/v3/job/details"
        querystring = {"myId": job_number}
        headers = {
            'authorization': "Basic *****************************=",
            'accept': "text/json",
            'content-type': "application/json",
            'cache-control': "no-cache"
        }
        response = requests.get(url, headers=headers, params=querystring)
        data = json.loads(response.text)
        
        if "num" not in data:
            # Generate a random 5-digit number
            job_num = str(random.randint(10000, 99999))
        else:
            job_num = data["num"]

        if response.status_code == 200:
            return True  # Equipment exists in Synchroteam
        else:
            return False  # Equipment does not exist in Synchroteam or an error occurred
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")

def create_customer(customer_name, customer_address):
    global Customer_ID
    global customer_Myid
    
    try:
        url = 'https://apis.synchroteam.com/api/v3/customer/send'
        headers = {
            'authorization': "Basic *********************************=",
            'accept': "text/json",
            'content-type': "application/json",
            'cache-control': "no-cache"
        }
        payload = {  
            "myId": customer_Myid,
            "name": customer_name,
            "address": customer_address
        }
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            return True  # Customer created successfully in Synchroteam
        else:
            return False  # Error occurred while creating the customer in Synchroteam
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")
        

def create_site(site_name, site_address, customer_id, site_id):
    global site_ID
    
    try:
        url = 'https://apis.synchroteam.com/api/v3/site/send'
        headers = {
            'authorization': "Basic ***************************************=",
            'accept': "text/json",
            'content-type': "application/json",
            'cache-control': "no-cache"
        }
        payload = {
            
            "myId": site_id,
            "name": site_name,
            "address": site_address,
            "customer": {
                "id": customer_id
            }
        }
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            return True  # Site created successfully in Synchroteam
        else:
            return False  # Error occurred while creating the site in Synchroteam
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")

def create_equipment(equipment_name, customer_id):
    try:
        global site_ID
        url = 'https://apis.synchroteam.com/api/v3/equipment/send'
        headers = {
            'authorization': "Basic *********************************=",
            'accept': "text/json",
            'content-type': "application/json",
            'cache-control': "no-cache"
        }
        payload = {
            "name": equipment_name,
            "site": {
                "id": site_ID
            },
            "customer": {
                "id": customer_id
            }
        }
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            return True  # Equipment created successfully in Synchroteam
        else:
            return False  # Error occurred while creating the equipment in Synchroteam
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")

def create_job(customer_name, site_name, site_id, equipment_name, job_description, job_number, site_address, report_data):
    try:
        global job_num
        global site_ID
        global equip_ID
        global stock_code
        global stock_description
        global stock_quantity
        global stock_status
        global stock_serial

        # Delete the previous job before creating a new one
        url = 'https://apis.synchroteam.com/api/v3/job/delete'
        headers = {
                'authorization': "Basic *************************************=",
                'accept': "text/json",
                'content-type': "application/json",
                'cache-control': "no-cache"
        }
        payload = {
                "myId": job_number,
        }
        response = requests.delete(url, headers=headers, json=payload)

        time.sleep(15)
        
        # Check if the customer, site, and equipment exist in Synchroteam
        customer_id = check_customer_exists(customer_name)
        site_exists = check_site_exists(site_name)
        equipment_exists = check_equipment_exists(equipment_name)

        if customer_id and site_exists and equipment_exists:
            url = 'https://apis.synchroteam.com/api/v3/job/send'
            headers = {
                'authorization': "Basic *************************************=",
                'accept': "text/json",
                'content-type': "application/json",
                'cache-control': "no-cache"
            }
            payload = {
                "myId": job_number,
                "num": job_num,
                "description": job_description,
                "customer": {
                    "id": customer_id,
                },
                "site": {
                    "id": site_ID,
                    "myId": site_id,
                    "name": site_name
                },
                "equipment": {
                    "id": equip_ID
                },
                "address": site_address,
                "reportTemplate": {
                    "id": 115104,
                    "name": "BCE Job Order V3",
                },
                "report": report_data
            }    
                       
            response = requests.post(url, headers=headers, json=payload)
            if response.status_code == 200:
                return True  # Job created or updated successfully in Synchroteam
            else:
                return False  # Error occurred while creating or updating the job in Synchroteam
        else:
            return False  # Customer, site, or equipment does not exist in Synchroteam
        
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")
        
def Send_email():
    global error_message
    # Account credentials
    username = "webhook@cvts.co.za"
    password = "*******************"

    # Email server details
    smtp_server = 'mail.your-server.de'  # Replace with your SMTP server
    smtp_port = 587  # Replace with your SMTP port

    # Email content
    from_addr = "webhook@cvts.co.za"
    to_addr = "sagenigeria@cvts.co.za"
    subject = "BCE NIGERIA SAGE INTERGRATION ERROR."
    body = error_message

    # Create the email header
    msg = MIMEMultipart()
    msg["From"] = from_addr
    msg["To"] = to_addr
    msg["Subject"] = subject

    # Attach the email body
    msg.attach(MIMEText(body, "plain"))

    # Connect to the server and send the email
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(username, password)
        server.sendmail(from_addr, to_addr, msg.as_string())
        print("")
    except Exception as e:
        print("")
    finally:
        server.quit()


server = '192.168.1.5'
database = 'BCE Lagos'
username = 'CVTS'
password = '*******************'

while True:
    try:
        # Create a new connection for each iteration
        conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
        connection = pyodbc.connect(conn_str)

        # Create a cursor object to execute SQL queries
        cursor = connection.cursor()

        # Execute the stored procedure
        cursor.execute("{CALL [dbo].[spGetSynchroteamData]}")

        # Get the column names from the cursor description
        column_names = [column[0] for column in cursor.description]

        # Fetch all the result rows
        result = cursor.fetchall()

        # Execute the second stored procedure [dbo].[spGetSynchroteamLineData]
        cursor.execute("{CALL [dbo].[spGetSynchroteamLineData]}")

        # Get the column names from the cursor description for the second result
        column_names_line = [column[0] for column in cursor.description]

        # Fetch all the result rows for the second result
        result_line_data = cursor.fetchall()

        # Load the workbook and sheet
        file_path = "C:\\Users\\CVTS\\Documents\\BCE NIGERIA SAGE INTEGRATION\\BCE NIGERIA.xlsx" 
        if not os.path.exists(file_path):
            # Create the file if it doesn't exist
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Initialize headers if the file is created anew
            sheet.append(["jobNumber", "dateModified"])
            workbook.save(file_path)
        else:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            # Initialize headers if the sheet is empty
            if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
                sheet.append(["jobNumber", "dateModified"])

        # Load existing data into a dictionary for quick access
        existing_data = {}
        if sheet.max_row > 1:  # Ensure there's more than just the header row
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                try:
                    job_number, date_modified = row
                    if isinstance(date_modified, str):
                        date_modified = datetime.strptime(date_modified, '%Y-%m-%d %H:%M:%S.%f')
                    existing_data[job_number] = date_modified
                except ValueError as ve:
                    #print(f"Error unpacking row: {row}, Error: {ve}")
                    continue

        # Process the result
        for row in result:
            row_id = row[column_names.index('dateModified')]
            job_number = str(row[column_names.index('jobNumber')])

            # Ensure row_id is a datetime object
            if isinstance(row_id, str):
                row_id = datetime.strptime(row_id, '%Y-%m-%d %H:%M:%S.%f')
            elif not isinstance(row_id, datetime):
                print(f"Unexpected type for row_id: {type(row_id)}")
                continue

            # Check if the row has already been processed
            if job_number in existing_data:
                existing_date = existing_data[job_number]
                if row_id <= existing_date:
                    continue  # Skip to the next row

            # Add the new row to the dictionary
            existing_data[job_number] = row_id

            # Continue with processing as in your original code
            customer_name = str(row[column_names.index('Customer Name')])
            customer_address = str(row[column_names.index('Customer Address')])
            customer_Myid = str(row[column_names.index('Customer ID')])
            customer_id = check_customer_exists(customer_name)

            customer_created_in_synchroteam = create_customer(customer_name, customer_address)
            if customer_created_in_synchroteam:
                customer_id = check_customer_exists(customer_name)
            else:
                error_message = "The Customer for this job " + job_number + " was not created successfully in Synchroteam"
                Send_email()

            site_name = str(row[column_names.index('Site Name')])
            site_id = str(row[column_names.index('Site ID')])
            check_site_exists(site_name)
            site_address = str(row[column_names.index('Site Address')])

            site_created_in_synchroteam = create_site(site_name, site_address, customer_id, site_id)
            if site_created_in_synchroteam:
                check_site_exists(site_name)
            else:
                error_message = "The Site for this job " + job_number + " was not created successfully in Synchroteam"
                Send_email()

            equipment_name = str(row[column_names.index('Equipment Name')])
            check_equipment_exists(equipment_name)
            equipment_created_in_synchroteam = create_equipment(equipment_name, customer_id)
            if not equipment_created_in_synchroteam:
                error_message = "The Equipment for this job " + job_number + " was not created successfully in Synchroteam"
                Send_email()

            job_description = str(row[column_names.index('Job Description')])
            site_address = str(row[column_names.index('Site Address')])
            check_if_job_exists(job_number)

            report_data_list = []
            iteration = 0

            for line in result_line_data:
                if line[1] == job_number:
                    filtered_row = line

                    # Unpack the values and assign them to the variables
                    _, _, stock_code, stock_description, stock_quantity, stock_serial, stock_status = line

                    Stock_Ref = None

                    if stock_status == 0:
                        Stock_Ref = "Quote"

                    if stock_status == 1:
                        Stock_Ref = "Active"

                    # Create a list of report data dictionaries for each iteration
                    report_data = {
                        "nmCategory": "Spares",
                        "nmItem": "Item Code",
                        "value": stock_code,
                        "Iteration": iteration
                    }
                    report_data_list.append(report_data)

                    report_data = {
                        "nmCategory": "Spares",
                        "nmItem": "Item Description",
                        "value": stock_description,
                        "Iteration": iteration
                    }
                    report_data_list.append(report_data)

                    report_data = {
                        "nmCategory": "Spares",
                        "nmItem": "Item Quantity",
                        "value": stock_quantity,
                        "Iteration": iteration
                    }
                    report_data_list.append(report_data)

                    report_data = {
                        "nmCategory": "Spares",
                        "nmItem": "Item Serial Number",
                        "value": stock_serial,
                        "Iteration": iteration
                    }
                    report_data_list.append(report_data)

                    report_data = {
                        "nmCategory": "Spares",
                        "nmItem": "Item Status",
                        "value": Stock_Ref,
                        "Iteration": iteration
                    }
                    report_data_list.append(report_data)
                    iteration += 1

            created_in_synchroteam = create_job(customer_name, site_name, site_id, equipment_name, job_description, job_number, site_address, report_data_list)

            if created_in_synchroteam:
                pass
            else:
                error_message = "There is a possible error in the Sage Integration code or database. job " + job_number + " was neither Created nor Updated"
                Send_email()

        # Delete rows older than 60 days and write back to Excel
        rows_to_keep = []
        for job_number, date_modified in existing_data.items():
            if datetime.now() - date_modified <= timedelta(days=60):
                rows_to_keep.append([job_number, date_modified.strftime('%Y-%m-%d %H:%M:%S.%f')])

        # Clear the sheet but keep the header
        sheet.delete_rows(2, sheet.max_row)

        # Add the rows to keep back to the sheet
        for row in rows_to_keep:
            sheet.append(row)

        # Save the workbook
        workbook.save(file_path)

        # Close the cursor and connection
        cursor.close()
        connection.close()

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        continue
    

